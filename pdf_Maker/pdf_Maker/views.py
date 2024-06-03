#PYTHON MODULE
import random
import img2pdf
from docx2pdf import convert as docx_to_pdf_convert



#SYSTEM 
import os
from django.conf import settings
from tempfile import NamedTemporaryFile

#DJANGO MODULES


from django.urls import reverse
from django.contrib import messages
from django.http import HttpResponse
from django.utils.html import strip_tags
from django.shortcuts import render,redirect
from django.core.files.base import ContentFile
from django.template.loader import render_to_string
from django.core.mail import send_mail,EmailMultiAlternatives
from django.contrib.auth.hashers import make_password, check_password



#MODELS
from user.models import UserInformation,UserFileHistory




#FILE OPERATION MODULES

from docx import Document
import mammoth
import pdfkit
from io import BytesIO
from PIL import Image, ExifTags
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph  





#--------------------------------home page--------------------------------------------------#

def home_view(request):
    
    return render(request,'index.html')
  

#-----------------user login and sign up and forget login credential-------------------------#
def otp_verification_mail(name,otpValue,validity,mail_to):
    context = {
         'user_name': name,
         'otp_code': otpValue,
          'valid_minutes': validity,  
        }
    html_message = render_to_string('mail_otp_to_user.html', context)
    plain_message = strip_tags(html_message)
    from_email = 'PDF Maker <s5tech.sendmail@gmail.com>'
    to = mail_to
    # Send the OTP email
    email_message = EmailMultiAlternatives(
        'Verify Your PDF Maker Account', plain_message, from_email, [to]
        )
    email_message.attach_alternative(html_message, "text/html")
    email_message.send()
    
    
      


def login_view(request):
    message_password_updated=request.GET.get('passwordUpdated',False)
    
    if request.method == 'POST':
        id=request.POST.get('id')
        password=request.POST.get('pass')
       
        try:
            user=UserInformation.objects.get(email=id)
        except UserInformation.DoesNotExist:
            try:
                user=UserInformation.objects.get(username=id)
            except UserInformation.DoesNotExist:
                user=None
        if user is not None and check_password(password,user.password):
                
            request.session['user_id'] = user.id
            
            return redirect('home')
        else:
                
            return render(request,'login.html',{'error':True})
     
    return render(request,'login.html',{'pass_updated':message_password_updated})
def logout_view(request):
    request.session.pop('user_id',None)
    return redirect('home')
def signup_view(request):
    if request.method =='POST':
        email=request.POST.get('uemail')
        
        #condition for checking if the account is already made by this email or not 
        hasAccount=UserInformation.objects.filter(email=email).first()
        if hasAccount is not None:
            return render(request,'login.html',{'error_alreadyAccountExist':True,'email':email})
        
        name=request.POST.get('uname')
        password=request.POST.get('upass')
       # password=make_password(sendPassword)
        
        #username creation
        username=""
        for char in email:
            if char == '@':
                break
            username+=char
        
        print("username from signup view  ",username)
        otpValue=""
        for i in range(0,6):
            otpValue+=str(random.randrange(0,9))
        
        #session data storation
        request.session['username']=username
        request.session['password']=password
        request.session['name']=name
        request.session['email']=email
        request.session['otp']=otpValue
        
       
       #sending otp to the associated mail
       #---------------outdated------------------------#
        # send_mail(
        #     'otp-verification','your otp is {}'.format(otpValue),
        #     's5tech.sendmail@gmail.com',[email],
        #      fail_silently=False,
        #   )
        #----------------------------------------------#
        
        otp_verification_mail(name,otpValue,3,email)
        
        
        
       
        
        return redirect('signup_otp')
    return render(request,'login.html')
def otp_view(request):
    name=request.session.get('name')
    email=request.session.get('email')
    username=request.session.get('username')
    if email is not None and username is not None:
        email = email.lower()
        username = username.lower()
    password=make_password(request.session.get('password'))  
    otp=request.session.get('otp')
    
 
    if request.method =='POST':
        user_otp=request.POST.get('otp')
        if user_otp == otp:
            
            user=UserInformation(email=email,password=password,name=name,username=username)
            user.save()
            request.session['user_id'] = user.id
            
            #pop or remove information from session
            request.session.pop('name',None)
            request.session.pop('username',None)
            request.session.pop('password',None)
            request.session.pop('email',None)
            request.session.pop('otp',None)
           
            
          
            return redirect('home')

        
        else:
            return render(request,'otp.html',{'incorrect':True})
    return render(request,'otp.html')       
def forget_password_view(request):
    if request.method == 'POST':
        # Get form data from the template
        email = request.POST.get('email')
        password = request.POST.get('password')
        repassword = request.POST.get('repassword')
        print("before try")
        try:
            print("inner try")
            
            # Attempt to retrieve user data based on the provided email
            user_data = UserInformation.objects.get(email=email)
            print("after user object retrieve")
            # Generate OTP for verification
            otpValue = ""
            for i in range(0, 6):
                otpValue += str(random.randrange(0, 9))

            # Store data in session for further verification
            request.session['email'] = email
            request.session['otp'] = otpValue
            request.session['password'] = make_password(password)

            # Sending mail to the user
            # send_mail(
            #     'otp-reset your password', 
            #     'your otp is {}'.format(otpValue),
            #     's5tech.sendmail@gmail.com', [email],
            #     fail_silently=False
            # )
            
            otp_verification_mail(user_data.name,otpValue,3,email)

            # Redirect to OTP verification page
            return redirect('forget_otp')
        except UserInformation.DoesNotExist:
            print("exeption")
            # User data doesn't exist for the provided email
            return render(request, 'forget_password.html', {'wrongEmail': True})
        except Exception as e:
            # Handle exception (e.g., log error, display error message)
            print("An error occurred while sending email:", e)

    # Render the forgetpass.html template for GET requests
    return render(request, 'forget_password.html')
def forget_otp_view(request):
    if request.method == 'POST':
        otp_from_user=request.POST.get('otp')
        otp_Backend_storage=request.session.get('otp')
        if otp_from_user == otp_Backend_storage:
            email=request.session.get('email')
            user=UserInformation.objects.get(email=email)
            updated_password=request.session.get('password')
            user.password=updated_password
            user.save()
            request.session['user_id'] = user.id
            request.session.pop('otp')
            request.session.pop('email')
            request.session.pop('password')
            
            
            return redirect('home')
        else:
            return render(request,'otp.html',{'incorrect':True})
            
    return render(request,'otp.html')



#--------------------------user account operations------------------------------------------#
def get_user(request):
    user_id=request.session.get('user_id')
    user=UserInformation.objects.get(pk=user_id)
    return user 
def profile_view(request):
    error=request.GET.get('errorOccurr',None)
    return render(request,'profile.html',{'error':error})
def update_picture_view(request):
    user=get_user(request)
    if request.method=='POST':
        if user.profile_picture:
            
            path_to_delete =os.path.join(settings.MEDIA_ROOT, str(user.profile_picture)) 
            if os.path.exists(path_to_delete):
                os.remove(path_to_delete)
        
        picture=request.FILES.get('profile_picture')
        user.profile_picture=picture
        user.save()
        
    return redirect('profile')
def remove_picture_view(request):
    user=get_user(request)
    if request.method=='POST':
        if user.profile_picture:
            path_to_delete =os.path.join(settings.MEDIA_ROOT, str(user.profile_picture))
            if os.path.exists(path_to_delete):
                os.remove(path_to_delete)
            user.profile_picture=None
            user.save()
    return redirect('profile')
def update_name_view(request):
    user=get_user(request)
    if request.method == 'POST':
        get_name=request.POST.get('name')
        if len(get_name) < 5:
            messages.error(request, 'Please enter atlease 5 characters.')
    
        else:
            user.name=get_name
            user.save()
    return redirect(reverse('profile') + f'?errorOccurr=name')
def update_username_view(request):
    user=get_user(request)
    if request.method == 'POST':
        get_username=request.POST.get('username')
        if len(get_username) < 5:
            messages.error(request, 'Please enter atlease 5 characters.')
            return redirect(reverse('profile') + f'?errorOccurr=username')  
        
        elif ' ' in get_username:
            messages.error(request, 'Username should not contain any spaces.')
            return redirect(reverse('profile') + f'?errorOccurr=username') 
    
        else:
            user.username=get_username
            user.save()
            
    
    return redirect('profile')
def delete_account_view(request):
    user=get_user(request)
    if request.method=='POST':
        password=request.POST.get('password')
        if user is not None and check_password(password,user.password):
            user.delete()           
            return redirect('logout')
            
        else:
            return render(request,'delete_account.html',{'userPassword':True})
    return render(request,'delete_account.html')
def user_history_view(request):
    user_id = request.session.get('user_id', None)
    if user_id is not None:
        user_files = UserFileHistory.objects.filter(user_id=user_id)
        return render(request, 'user_history.html', {'user_files': user_files})
    
    return render(request,'user_history.html')


#-----------------------------file conversion-----------------------------------------#

def isActive(request):
    user = request.session.get('user_id', None)
    if user is not None:
        return True
    return False
def store_user_history(request, user_filename, user_file_content, pdf_filename, pdf_content):
    user_id=request.session.get('user_id')
    if user_id:
        user=UserInformation.objects.get(pk=user_id)
        user_file=ContentFile(user_file_content.getvalue(),name=user_filename)
        pdf_file=ContentFile(pdf_content,name=pdf_filename)
        UserFileHistory.objects.create(
            user=user,
            user_file=user_file,
            pdf_file=pdf_file
        )
     
def text_to_pdf_view(request):
    if request.method == 'POST':
        user_file = request.FILES.get('file')
        user_filename=user_file.name
        user_file_content = BytesIO()
        for chunk in user_file.chunks():
            user_file_content.write(chunk)     
        if user_file:
            size_limit_kb = 600
            size_limit_bytes = size_limit_kb * 1024
            if user_file.size > size_limit_bytes:
                return render(request, 'pdf.html', {'file_accept': '.txt', 'file_size_exceeded': True})
            pdf_content = convert_text_to_pdf(user_file)
            pdf_filename=user_file.name.replace('.txt','.pdf')
            if isActive(request):
                store_user_history(request,user_filename,user_file_content,pdf_filename,pdf_content)
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{user_file.name.replace(".txt", ".pdf")}"'
            return response
    return render(request, 'file_converter.html', {'file_accept': '.txt'})
def convert_text_to_pdf(file):
    pdf_buffer = BytesIO()
    pdf = SimpleDocTemplate(pdf_buffer, pagesize=letter)    
    styles = getSampleStyleSheet()    
    story = []
    if file.name.endswith('.txt'):
        with file.open(mode='r') as txt_content:
            for line in txt_content:
                story.append(Paragraph(line, styles['Normal']))       
    pdf.build(story)
    pdf_buffer.seek(0)
    return pdf_buffer.getvalue()
def img_to_pdf_view(request):
    if request.method == 'POST':
        user_file = request.FILES.get('file')       
        if user_file:
            user_filename = user_file.name            
            user_file_content = BytesIO()
            for chunk in user_file.chunks():
                user_file_content.write(chunk)             
            pdf_content = convert_image_to_pdf(user_file_content)           
            pdf_filename = user_filename.rsplit('.', 1)[0] + '.pdf'           
            if isActive(request):
                store_user_history(request, user_filename, user_file_content, pdf_filename, pdf_content)           
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{pdf_filename}"'
            return response
    return render(request, 'file_converter.html', {'file_accept': '.png, .jpg, .jpeg'})
def convert_image_to_pdf(image_file):
    """
    Converts an image file to PDF format.
    
    Parameters:
    image_file (BytesIO): A file-like object containing the image to be converted.
    
    Returns:
    bytes: The PDF content as a byte string.
    
    img2pdf.ExifOrientationError: Raised when img2pdf encounters invalid rotation
    information in the Exif metadata of the image.
    """
    
    image_file.seek(0) 
    
    try:
        pdf_content = img2pdf.convert(image_file)
        return pdf_content
    except img2pdf.ExifOrientationError:
        # Handle ExifOrientationError by correcting the image rotation
        image_file.seek(0)
        image = Image.open(image_file)
        
        # Fixing the orientation based on Exif data
        try:
            for orientation in ExifTags.TAGS.keys():
                if ExifTags.TAGS[orientation] == 'Orientation':
                    break
            exif = dict(image._getexif().items())
            
            if exif[orientation] == 3:
                image = image.rotate(180, expand=True)
            elif exif[orientation] == 6:
                image = image.rotate(270, expand=True)
            elif exif[orientation] == 8:
                image = image.rotate(90, expand=True)
        except (AttributeError, KeyError, IndexError):
            # No Exif data or no orientation info
            pass
        
        # Save the corrected image to a BytesIO object
        corrected_image_content = BytesIO()
        image.save(corrected_image_content, format=image.format)
        corrected_image_content.seek(0)
        
        # Convert the corrected image content to a PDF
        pdf_content = img2pdf.convert(corrected_image_content)
        return pdf_content



def docx_to_pdf_view(request):
    if request.method == 'POST':
        user_file = request.FILES.get('file')
        if user_file:
            user_filename = user_file.name
            
            # Read the uploaded .docx file into a BytesIO object
            docx_file_content = BytesIO()
            for chunk in user_file.chunks():
                docx_file_content.write(chunk)
            docx_file_content.seek(0)  # Reset file pointer to the beginning
            
            # Convert the .docx file to PDF
            pdf_file_content = convert_docx_to_pdf(docx_file_content)
            
            if isActive(request):
                pdf_filename = os.path.splitext(user_filename)[0] + '.pdf'
                store_user_history(request, user_filename, docx_file_content, pdf_filename, pdf_file_content.getvalue())
            
            # Create HTTP response with the PDF file
            response = HttpResponse(pdf_file_content, content_type='application/pdf')
            pdf_filename = os.path.splitext(user_filename)[0] + '.pdf'
            response['Content-Disposition'] = f'attachment; filename="{pdf_filename}"'
            return response
    
    return render(request, 'file_converter.html', {'file_accept': '.docx'})

def convert_docx_to_pdf(docx_file_content):
    # Use NamedTemporaryFile to create a temporary .docx file
    with NamedTemporaryFile(delete=False, suffix='.docx') as temp_docx:
        temp_docx.write(docx_file_content.read())
        temp_docx.flush()
        temp_docx_path = temp_docx.name
    
    # Create another temporary file for the PDF output
    with NamedTemporaryFile(delete=False, suffix='.pdf') as temp_pdf:
        temp_pdf_path = temp_pdf.name
    
    # Convert .docx to PDF using the paths of the temporary files
    docx_to_pdf_convert(temp_docx_path, temp_pdf_path)
    
    # Read the converted PDF into a BytesIO object
    pdf_file_content = BytesIO()
    with open(temp_pdf_path, 'rb') as pdf_file:
        pdf_file_content.write(pdf_file.read())
    pdf_file_content.seek(0)  # Reset file pointer to the beginning
    
    return pdf_file_content

def excel_to_pdf_view(request):
    if request.method == 'POST':
        user_file = request.FILES.get('file')
        if user_file:
            pdf_filename=user_file.name.replace('.xlsx', '.pdf')
            pdf_content=convert_excel_to_pdf(user_file)
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{pdf_filename}"'
            return response
    return render(request, 'file_converter.html', {'file_accept': '.xlsx'})
            
        
def convert_excel_to_pdf(excel_file):
    # Load the Excel file
    wb = load_workbook(excel_file)

    # Select the active worksheet
    ws = wb.active

    # Create a BytesIO object to hold the PDF data
    pdf_buffer = BytesIO()

    # Create a canvas for PDF generation
    c = canvas.Canvas(pdf_buffer, pagesize=letter)

    # Iterate through rows and columns to extract data
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            c.drawString(100, 700, str(cell))  # Adjust position as needed

    # Save the PDF content
    c.save()

    # Reset buffer position to the beginning
    pdf_buffer.seek(0)

    return pdf_buffer.getvalue()







def compress_image_view(request):
    if request.method == 'POST':
        image_file = request.FILES.get('file')
        
        target_size= int(request.POST.get('target_size')) 
        unit=request.POST.get('unit')
        
        if unit == 'MB':
            # converting the mb to kb
            target_size= target_size * 1024

        
        # Call the function for compression
        compressed_img = compress_image(image_file, target_size)
        
        # Prepare response
        response = HttpResponse(compressed_img.getvalue(), content_type='image/jpeg')
        response['Content-Disposition'] = 'attachment; filename="compressed_image.jpg"'
        return response

    return render(request, 'compress_file.html', {'file_type': '.png, .jpg, .jpeg'})
def compress_image(file, target_size_kb=300):
    img = Image.open(file)
    original_size = img.size[0] * img.size[1] * 3  
    
    
    # Initialize quality parameters
    min_quality = 0
    max_quality = 100
    quality = (min_quality + max_quality) // 2 
    
    # Compress the image using binary search
    while True:
        # Compress the image
        compressed_img = BytesIO()
        img.save(compressed_img, format='JPEG', quality=quality)
        
        # Get the size of the compressed image
        compressed_size = len(compressed_img.getvalue())
        
        
        # Check if the compressed size is within the target range
        if compressed_size <= target_size_kb * 1024:
            break  # If compressed size is within target, break the loop
        else:
            # Adjust quality based on binary search
            if compressed_size > target_size_kb * 1024:
                max_quality = quality - 1
            else:
                min_quality = quality + 1
            
            # Check if quality becomes too low or too high
            if min_quality > max_quality:
                # Compression quality too low, break the loop
                break
            
            # Update quality for next iteration
            quality = (min_quality + max_quality) // 2
    
    return compressed_img
def compress_pdf_view(request):
    if request.method == 'POST':
        pdf_file = request.FILES['file']
        target_size= int(request.POST.get('target_size')) 
        unit=request.POST.get('unit')
        
        if unit == 'MB':
            # converting the mb to kb
            target_size *= 1024
        
        # Call the function for compression
        compressed_pdf = compress_pdf(pdf_file,target_size)
        
        # Prepare response
        response = HttpResponse(compressed_pdf.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="compressed_pdf.pdf"'
        return response
    return render(request,'compress_file.html',{'file_type':'.pdf'})






#----------------------------Feedback------------------------------------------#

def feedback_view(request):
    if request.method == 'POST':
        name=request.POST.get('name')
        print('name')
        email=request.POST.get('email')
       
        APP='PDF_Maker'
        admin_gmail='feedback.s5tech@gmail.com'
        feedback=request.POST.get('feedback')
        message=request.POST.get('message')
        
        send_mail(
            '{}-feedback'.format(APP),
            'Name: {}\n\nEmail:{}\n\nfeedback_type: {}\nSubject\n\n\{}'.format(name,email,feedback,message),
            's5tech.sendmail@gmail.com',
            [admin_gmail],
            fail_silently=False,
           
            
        )
        return redirect('home')
    return render(request,'feedback.html')

def privacy_policy_view(request):
    return render(request,'privacy_policy.html')
def terms_and_conditions_view(request):
    return render(request,'terms_and_conditions.html')